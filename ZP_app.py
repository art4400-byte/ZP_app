import tkinter as tk
from tkinter import messagebox, filedialog
from openpyxl import load_workbook

# Константы
sum_night = 186.69
sum_day = 162.34
additional_amount_350 = 350  # Фиксированная сумма для добавления
additional_amount_not_fine = 66  # Фиксированная сумма для добавления
additional_amount_750 = 750  # Фиксированная сумма для добавления
# Глобальная переменная для хранения результата
result = 0
medium_zp = ((((3200 + 3500 + 3100 + 2500 + 3300 + 3200) / 6) / 30) - 20.9)


# Функция для создания окна результата теста
def result_test_window():
    test_window = tk.Toplevel()
    test_window.title("Ввод результата теста")
    test_window.geometry("300x150")
    test_window.grab_set()  # Делаем окно модальным

    tk.Label(test_window, text="Введите результат теста (0-100):").pack(pady=10)

    # Создаем строку для ввода реузльтата теста
    test_entry = tk.Entry(test_window)
    test_entry.pack(pady=5)

    # Функция для проверки типа введеных данных и удаления окна результата теста
    def apply_test_result():
        try:
            test_result = float(test_entry.get())
            if 0 <= test_result <= 100:
                update_constants(test_result)  # Добовляем результат в функцию
                test_window.destroy()  # Удаляем окно результата теста
                root.deiconify()  # Показываем основное окно "Подсчет ЗП"
            else:
                # Проверка на промежуток от 0 до 100
                messagebox.showerror("Ошибка", "Введите число от 0 до 100")

        except ValueError:
            # Проверка на тип дынных введеного значения
            messagebox.showerror("Ошибка", "Пожалуйста, введите число")

    # Создание кнопки "Применить"
    tk.Button(test_window, text="Применить", command=apply_test_result).pack(pady=10)


# Функция для корректирования константы в зависимости от результата теста
def update_constants(test_result):
    global sum_day, sum_night

    if 0 <= test_result <= 86:
        sum_night = 166.69
        sum_day = 142.34
    elif 86 < test_result <= 97:
        sum_night = 176.69
        sum_day = 152.34
    else:  # test_result > 98:
        sum_night = 186.69
        sum_day = 162.34


# Создаем основное окно
root = tk.Tk()
root.title("Подсчет ЗП")
root.geometry("600x400")  # Создаем основное окно
root.withdraw()  # Скрываем основное окно

# Показываем окно ввода результата теста перед основным окном
result_test_window()


# Функция для вычисления результата
def calculate():
    global result
    try:
        # Получаем значения из полей ввода
        day = float(entry_day.get())
        night = float(entry_night.get())
        subtract = float(entry_subtract.get())
        c_day = float(entry_off_day.get())
        c_night = float(entry_off_night.get())
        double_day = float(entry_double_day.get())
        double_night = float(entry_double_night.get())
        extra_day = float(entry_extra_day.get())
        sick = float(entry_sick.get())
        taxi = float(entry_taxi.get())

        # Проверка на отрицательные значения
        if day < 0 or night < 0 or c_day < 0 or c_night < 0:
            messagebox.showerror("Ошибка", "Количество смен и ДО не может быть отрицательным.")
            return

        # Вычисляем результат
        result = ((((day - double_day) * sum_day) + ((night - double_night - extra_day) * sum_night))
                  + (((sum_day * 2) * double_day) + ((sum_night * 2) * double_night))
                  + ((c_day * sum_day) + (c_night * sum_night))
                  + ((sum_night * 1.2) * extra_day)
                  + (medium_zp * sick)
                  + taxi - subtract - 100)
        update_result_label()

    except ValueError:
        messagebox.showerror("Ошибка", "Пожалуйста, введите корректные числа.")


# Создаем функцию для добавления данных в .xlxs
def export_to_excel():
    if result == 0:
        messagebox.showwarning("Предупреждение", "Сначала выполните расчеь!")
        return

    try:
        # Возможность выбора файла для сохранения
        filepath = filedialog.askopenfilename(
            filetypes=[("Excel файлы", "*.xlsx"), ("Все файлы", "*.*")],
            title="Выберите файл Excel для сохранения"
        )

        if not filepath:    # Если пользоваетль отменил выбор
            return

        # Загрузка существующего файла .xlxs
        wb = load_workbook(filepath)

        # Выбираем активный лист
        ws = wb.active

        # Находит первую свободную строку для записи в указаном столбце
        target_column = "B" # Указывается нужный столбец
        row_num = 1
        while ws[f"{target_column}{row_num}"].value is not None:
            row_num += 1

        # Запись данных
        ws[f"{target_column}{row_num}"] = result

        # Сохранение изменеий
        wb.save(filepath)
        messagebox.showinfo("Успех", f"Данные добавлены в файл:\n{filepath}")

    except Exception as e:
        messagebox.showerror("Ошибка", f"Не удалось сохаранить данные: \n{str(e)}")


# Функция для добавления 350 к результату
def add_350():
    global result
    result += additional_amount_350
    update_result_label()


# Функция для добавления 750 к результату
def add_750():
    global result
    result += additional_amount_750
    update_result_label()


# Функция для добавления 66 к результату
def add_not_fine():
    global result
    result += additional_amount_not_fine
    update_result_label()


# Функция для обновления отображения результата
def update_result_label():
    result_label.config(text=f"Текущий результат: {result:.2f}")


# Функция для очистки полей ввода
def clear_entries():
    entry_night.delete(0, tk.END)
    entry_day.delete(0, tk.END)
    entry_subtract.delete(0, tk.END)
    entry_off_day.delete(0, tk.END)
    entry_off_night.delete(0, tk.END)
    entry_double_day.delete(0, tk.END)
    entry_double_night.delete(0, tk.END)
    entry_extra_day.delete(0, tk.END)
    entry_sick.delete(0, tk.END)
    entry_taxi.delete(0, tk.END)
    global result
    result = 0
    update_result_label()


# Используем grid для расположения элементов

# Заголовки
label_night = tk.Label(root, text="Ночные смены")
label_night.grid(row=0, column=0, padx=10, pady=5)

label_day = tk.Label(root, text="Дневные смены")
label_day.grid(row=0, column=1, padx=10, pady=5)

label_day = tk.Label(root, text="Дополнительное")
label_day.grid(row=0, column=2, padx=10, pady=5)

# Поле для ввода количества ночных смен
label_night_input = tk.Label(root, text="Кол-во ночных смен:")
label_night_input.grid(row=1, column=0, padx=10, pady=5)
entry_night = tk.Entry(root, width=20)
entry_night.grid(row=2, column=0, padx=10, pady=5)

# Поле для ввода количества дневных смен
label_day_input = tk.Label(root, text="Кол-во дневных смен:")
label_day_input.grid(row=1, column=1, padx=10, pady=5)
entry_day = tk.Entry(root, width=20)
entry_day.grid(row=2, column=1, padx=10, pady=5)

# Поле для ввода вычет
label_subtract_input = tk.Label(root, text="Сумма вычет:")
label_subtract_input.grid(row=1, column=2, padx=10, pady=5)
entry_subtract = tk.Entry(root, width=20)
entry_subtract.grid(row=2, column=2, padx=10, pady=5)

# Поле для ввода отчета за такси
label_taxi = tk.Label(root, text="Отчет за такси:")
label_taxi.grid(row=3, column=2, padx=10, pady=5)
entry_taxi = tk.Entry(root, width=20)
entry_taxi.grid(row=4, column=2, padx=10, pady=5)

# Поле для ввода количества ДО ночных
label_off_night = tk.Label(root, text="Кол-во ночных ДО:")
label_off_night.grid(row=3, column=0, padx=10, pady=5)
entry_off_night = tk.Entry(root, width=20)
entry_off_night.grid(row=4, column=0, padx=10, pady=5)

# Поле для ввода количества ДО дневных
label_off_day = tk.Label(root, text="Кол-во дневных ДО:")
label_off_day.grid(row=3, column=1, padx=10, pady=5)
entry_off_day = tk.Entry(root, width=20)
entry_off_day.grid(row=4, column=1, padx=10, pady=5)

# Поле для ввода количества рабочих ночных в праздничные
label_double_night = tk.Label(root, text="Ночные смены в праздники:")
label_double_night.grid(row=5, column=0, padx=10, pady=5)
entry_double_night = tk.Entry(root, width=20)
entry_double_night.grid(row=6, column=0, padx=10, pady=5)

# Поле для ввода количества рабочих дневных в праздничные
label_double_day = tk.Label(root, text="Дневные смены в праздники:")
label_double_day.grid(row=5, column=1, padx=10, pady=5)
entry_double_day = tk.Entry(root, width=20)
entry_double_day.grid(row=6, column=1, padx=10, pady=5)

# Поле для ввода количества смен выходного дня
label_extra_day = tk.Label(root, text="Смена выходного дня:")
label_extra_day.grid(row=7, column=0, padx=10, pady=5)
entry_extra_day = tk.Entry(root, width=20)
entry_extra_day.grid(row=8, column=0, padx=10, pady=5)

# Поле для ввода количества дней по больничному
label_sick = tk.Label(root, text="Дней по больничному")
label_sick.grid(row=7, column=1, padx=10, pady=5)
entry_sick = tk.Entry(root, width=20)
entry_sick.grid(row=8, column=1, padx=10, pady=5)

# Кнопка для экспорта в Excel
button_export = tk.Button(root, text="Экспорт в Excel", command=export_to_excel)
button_export.place(relx=0.74, rely=0.92, anchor="center")

# Кнопка для вычисления
button_calculate = tk.Button(root, text="Вычислить", command=calculate)
button_calculate.place(relx=0.66, rely=0.78, anchor="center")

# Кнопка для очистки полей ввода
button_clear = tk.Button(root, text=" Очистить ", command=clear_entries)
button_clear.place(relx=0.82, rely=0.78, anchor="center")

# Кнопка для добавления 350
button_add_350 = tk.Button(root, text="Добавить 350", command=add_350)
button_add_350.grid(row=5, column=2, pady=10)

# Кнопка для добавления 750
button_add_750 = tk.Button(root, text="Добавить 750", command=add_750)
button_add_750.grid(row=6, column=2, pady=10)

# Кнопка для добавления премии без штрафов
button_add_not_fine = tk.Button(root, text="Премия без штрафов", command=add_not_fine)
button_add_not_fine.grid(row=7, column=2, pady=10)

# Метка для отображения текущего результата
result_label = tk.Label(root, text="Текущий результат: 0.00")
result_label.grid(row=9, column=2, columnspan=2, pady=10)

# Запуск основного цикла
root.mainloop()

# Тест 1.2 проверка